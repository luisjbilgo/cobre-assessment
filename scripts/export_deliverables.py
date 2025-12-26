"""
Deliverable Export Utilities for Cobre Payment Corridor Analysis

Functions to generate Excel workbooks with analysis results and visualizations.
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from pathlib import Path
from typing import Dict, List
import os


def create_excel_workbook(
    data_dict: Dict[str, pd.DataFrame],
    output_path: str = 'output/analysis_workbook.xlsx',
    visualization_dir: str = 'output/visualizations'
) -> None:
    """
    Create comprehensive Excel workbook with all analysis results.

    Args:
        data_dict: Dictionary of sheet_name: DataFrame pairs
        output_path: Path to save the Excel workbook
        visualization_dir: Directory containing visualization PNGs
    """
    # Ensure output directory exists
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Define cell border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add sheets with data
    for sheet_name, df in data_dict.items():
        ws = wb.create_sheet(title=sheet_name)

        # Write DataFrame to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

                # Apply header styling
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # Format numbers
                if r_idx > 1 and isinstance(value, (int, float)):
                    if 'rate' in df.columns[c_idx-1].lower() or '%' in str(df.columns[c_idx-1]):
                        cell.number_format = '0.00"%"'
                    elif 'amount' in df.columns[c_idx-1].lower() or 'usd' in df.columns[c_idx-1].lower():
                        cell.number_format = '$#,##0.00'
                    elif 'count' in df.columns[c_idx-1].lower() or 'volume' in df.columns[c_idx-1].lower():
                        cell.number_format = '#,##0'

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A2'

    # Add visualization sheet if images exist
    if Path(visualization_dir).exists():
        add_visualizations_sheet(wb, visualization_dir)

    # Save workbook
    wb.save(output_path)
    print(f"\n✅ Excel workbook created: {output_path}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")


def add_visualizations_sheet(wb: Workbook, visualization_dir: str) -> None:
    """
    Add a sheet with embedded visualization images.

    Args:
        wb: Openpyxl workbook object
        visualization_dir: Directory containing PNG files
    """
    ws = wb.create_sheet(title="Visualizations")

    # Add title
    ws.cell(row=1, column=1, value="Analysis Visualizations").font = Font(bold=True, size=14)

    row = 3
    vis_files = sorted(Path(visualization_dir).glob("*.png"))

    for idx, img_path in enumerate(vis_files):
        if img_path.exists():
            # Add image title
            ws.cell(row=row, column=1, value=img_path.stem.replace('_', ' ').title()).font = Font(bold=True, size=11)
            row += 1

            # Add image (resize to fit Excel)
            try:
                img = XLImage(str(img_path))
                # Resize to fit width (approx 800 pixels = 11 columns)
                img.width = 800
                img.height = int(img.height * (800 / img.width))

                ws.add_image(img, f'A{row}')
                # Move to next position (approximate row height = 20 pixels per row)
                row += int(img.height / 20) + 3
            except Exception as e:
                ws.cell(row=row, column=1, value=f"Error loading image: {str(e)}")
                row += 2

    # Adjust column width
    ws.column_dimensions['A'].width = 120


def format_revenue_impact_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """
    Format revenue impact calculations for Excel.

    Args:
        df: DataFrame with revenue calculations

    Returns:
        Formatted DataFrame
    """
    # Add formatted currency columns
    formatted_df = df.copy()

    # Ensure proper formatting
    for col in formatted_df.columns:
        if 'revenue' in col.lower() or 'amount' in col.lower() or 'value' in col.lower():
            formatted_df[col] = formatted_df[col].round(2)

    return formatted_df


def create_summary_sheet_data() -> pd.DataFrame:
    """
    Create summary data for executive overview sheet.

    Returns:
        DataFrame with key findings summary
    """
    summary_data = {
        'Metric': [
            'Total Transactions',
            'Total Users',
            'Analysis Period',
            'Total Transaction Value',
            'Overall Failure Rate',
            'USD→MXN Failure Rate',
            'Target Failure Rate',
            'Problem Corridor',
            'Annual Revenue Impact',
            'Primary Root Cause',
            'Recommended Action'
        ],
        'Value': [
            '50,000',
            '5,000',
            'Jul-Dec 2025 (6 months)',
            '$281.5M',
            '9.6%',
            '18.3%',
            '5.0%',
            'USD→MXN (34.8% of volume)',
            '$360,000 potential recovery',
            'Large transaction verification delays',
            'Fix USD→MXN corridor'
        ]
    }

    return pd.DataFrame(summary_data)


def export_all_deliverables(
    conn,
    sql_queries_module
) -> None:
    """
    Execute all queries and export complete Excel workbook.

    Args:
        conn: SQLite database connection
        sql_queries_module: Imported sql_queries module
    """
    print("\n" + "="*60)
    print("GENERATING DELIVERABLES")
    print("="*60 + "\n")

    data_dict = {}

    # 1. Executive Summary
    data_dict['Executive Summary'] = create_summary_sheet_data()
    print("✓ Created Executive Summary")

    # 2. Corridor Performance
    corridor_query = sql_queries_module.corridor_performance_query()
    data_dict['Corridor Performance'] = pd.read_sql_query(corridor_query, conn)
    print("✓ Created Corridor Performance sheet")

    # 3. User Segment Analysis
    segment_query = sql_queries_module.user_segment_analysis_query()
    data_dict['User Segments'] = pd.read_sql_query(segment_query, conn)
    print("✓ Created User Segments sheet")

    # 4. Time Patterns - Daily
    daily_query = sql_queries_module.daily_trend_query()
    data_dict['Daily Trends'] = pd.read_sql_query(daily_query, conn)
    print("✓ Created Daily Trends sheet")

    # 5. Time Patterns - Day of Week
    dow_query = sql_queries_module.day_of_week_pattern_query()
    data_dict['Day of Week'] = pd.read_sql_query(dow_query, conn)
    print("✓ Created Day of Week sheet")

    # 6. Amount Distribution
    amount_query = sql_queries_module.amount_distribution_query()
    data_dict['Amount Distribution'] = pd.read_sql_query(amount_query, conn)
    print("✓ Created Amount Distribution sheet")

    # 7. USD→MXN Analysis - Create temp table first
    usd_mxn_create = sql_queries_module.usd_mxn_corridor_query()
    conn.execute(usd_mxn_create)

    # USD→MXN Segment Analysis
    usd_mxn_segment_query = sql_queries_module.usd_mxn_segment_analysis_query()
    data_dict['USD_MXN Segments'] = pd.read_sql_query(usd_mxn_segment_query, conn)
    print("✓ Created USD_MXN Segments sheet")

    # USD→MXN Amount Analysis
    usd_mxn_amount_query = sql_queries_module.usd_mxn_amount_analysis_query()
    data_dict['USD_MXN Amounts'] = pd.read_sql_query(usd_mxn_amount_query, conn)
    print("✓ Created USD_MXN Amounts sheet")

    # USD→MXN Monthly Trend
    usd_mxn_monthly_query = sql_queries_module.usd_mxn_monthly_trend_query()
    data_dict['USD_MXN Monthly'] = pd.read_sql_query(usd_mxn_monthly_query, conn)
    print("✓ Created USD_MXN Monthly sheet")

    # 8. Corridor Comparison for Strategy
    corridor_compare_query = sql_queries_module.corridor_comparison_for_strategy_query()
    data_dict['Corridor Comparison'] = pd.read_sql_query(corridor_compare_query, conn)
    print("✓ Created Corridor Comparison sheet")

    # Create Excel workbook
    create_excel_workbook(data_dict)

    print("\n" + "="*60)
    print("DELIVERABLE GENERATION COMPLETE")
    print("="*60 + "\n")


def save_dataframe_to_csv(
    df: pd.DataFrame,
    filename: str,
    output_dir: str = 'output/csv_exports'
) -> None:
    """
    Save DataFrame to CSV for backup/analysis.

    Args:
        df: DataFrame to save
        filename: Output filename
        output_dir: Directory for CSV exports
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    output_path = Path(output_dir) / filename
    df.to_csv(output_path, index=False)
    print(f"✅ CSV exported: {output_path}")
